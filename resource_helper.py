import os
import glob
import re

from openpyxl import Workbook
from openpyxl import load_workbook
import xlsxwriter

#Modify the 0.7 value to modify the Hundred KPI
kpiVal = 0.8

def kpiHund(siteBw, prevHun):

    if siteBw >= (1100*kpiVal):
        result = 12 - prevHun
    if siteBw >= (1000*kpiVal):
        result = 11 - prevHun
    if siteBw >= (900*kpiVal):
        result = 10 - prevHun
    elif siteBw >= (800*kpiVal):
        result = 9 - prevHun
    elif siteBw >= (700*kpiVal):
        result = 8 - prevHun
    elif siteBw >= (600*kpiVal):
        result = 7 - prevHun
    elif siteBw >= (500*kpiVal):
        result = 6 - prevHun
    elif siteBw >= (400*kpiVal):
        result = 5 - prevHun
    elif siteBw >= (300*kpiVal):
        result = 4 - prevHun
    elif siteBw >= (200*kpiVal):
        result = 3 - prevHun
    elif siteBw >= (100*kpiVal):
        result = 2 - prevHun
    else:
        result = 1 - prevHun

    return result

def kpiTen(siteBw, prevTen):
    if siteBw >= (30*kpiVal):
        result = 4 - prevTen
    elif siteBw >= (20*kpiVal):
        result = 3 - prevTen
    elif siteBw >= (10*kpiVal):
        result = 2 - prevTen
    else:
        result = 1 - prevTen

    if result == -1:
        result = 0
    return result

def matrixWriter(result, filename):

    wb = load_workbook(filename)
    ws = wb.active
    firstList = []
    for row in range (1,6):
        data = []
        for col in ws.iter_cols(min_col=1, min_row=row, max_col=39, max_row=row):
            for cell in col:
                data.append(cell.value)
        firstList.append(data)

    secondList = []
    for row in range (3,ws.max_row+1):
        data = []
        for col in ws.iter_cols(min_col=1, min_row=row, max_col=9, max_row=row):
            for cell in col:
                data.append(cell.value)
        secondList.append(data)


    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet()


    row = -1

    for i in range(len(firstList)):
        row = row+1
        col = -1
        for t in range(len(firstList[0])):
            col = col+1
            worksheet.write(row, col, firstList[i][t])


    row = 1

    for i in range(len(secondList)):
        row = row+1
        col = -1
        for t in range(len(secondList[0])):
            col = col+1
            worksheet.write(row, col, secondList[i][t])


    col = 8

    for i in range(len(result)):
        col = col+1
        row = 1
        for t in range(len(result[0])):
            row = row + 1
            worksheet.write(row, col, result[i][t])

    workbook.close()
    return True

def getSiteList(filename):
    wb = load_workbook(filename)
    ws = wb.active
    siteList = []
    for row in range (3,ws.max_row+1):
    # for row in range (3,6):
        data = []
        for col in ws.iter_cols(min_col=1, min_row=row, max_col=1, max_row=row):
            for cell in col:
                data.append(cell.value)

        siteList.append(data[0])

    return siteList

def getInitialbw(filename):
    wb = load_workbook(filename)
    ws = wb.active
    initialBw = []
    for row in range (3,ws.max_row+1):
    # for row in range (3,6):
        data = []
        for col in ws.iter_cols(min_col=6, min_row=row, max_col=6, max_row=row):
            for cell in col:
                data.append(cell.value)

        initialBw.append(round(data[0],2))

    return initialBw

def getTenDemands(filename):
    wb = load_workbook(filename)
    ws = wb.active
    TenDemands = []
    for row in range (3,ws.max_row+1):
    # for row in range (3,6):
        data = []
        for col in ws.iter_cols(min_col=8,min_row=row, max_col=8, max_row=row):
            for cell in col:
                data.append(cell.value)

        TenDemands.append(data[0])

    return TenDemands

def getHundDemands(filename):
    wb = load_workbook(filename)
    ws = wb.active
    HundDemands = []
    for row in range (3,ws.max_row+1):
    # for row in range (3,6):
        data = []
        for col in ws.iter_cols(min_col=9, min_row=row, max_col=9, max_row=row):
            for cell in col:
                data.append(cell.value)

        HundDemands.append(data[0])

    return HundDemands



def getSitebw(filename, siteList, site):
    siteBw = 0

    wb = load_workbook(filename)
    ws = wb.active

    siteList = []

    for row in range (3,ws.max_row+1):
        data = []
        for col in ws.iter_cols(min_row=row, max_col=5, max_row=row):
            for cell in col:
                data.append(cell.value)
        siteList.append(data)

    for i in range(len(siteList)):
        if site in siteList[i]:
            siteBw = siteBw + siteList[i][4]
            # siteBw = round(siteBw,2)

    return siteBw

# Copyrigth David Cobos: dacobos@cisco.com
